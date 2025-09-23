"""SafeWork 보고서 생성 시스템"""

import io
import pandas as pd
from datetime import datetime, timedelta, date
from flask import Blueprint, jsonify, request, send_file, render_template
from flask_login import login_required, current_user
from sqlalchemy import text, func, extract
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from models import db
from routes.admin import admin_required

safework_reports_bp = Blueprint("safework_reports", __name__)


@safework_reports_bp.route("/dashboard")
@admin_required
def reports_dashboard():
    """보고서 생성 대시보드"""
    return render_template("admin/safework_reports_dashboard.html")


@safework_reports_bp.route("/health-check-summary")
@admin_required
def health_check_summary_report():
    """건강검진 현황 종합 보고서 생성"""
    try:
        # 기본 매개변수
        year = request.args.get("year", datetime.now().year, type=int)
        format_type = request.args.get("format", "json")  # json, excel

        # 건강검진 통계 데이터
        report_data = {}

        # 1. 전체 통계
        total_workers_query = db.session.execute(
            text(
                """
            SELECT COUNT(*) as total_workers,
                   COUNT(CASE WHEN is_active = 1 THEN 1 END) as active_workers,
                   COUNT(CASE WHEN is_active = 0 THEN 1 END) as inactive_workers
            FROM safework_workers
        """
            )
        )
        total_stats = total_workers_query.fetchone()

        report_data["총근로자수"] = total_stats[0] if total_stats else 0
        report_data["재직근로자수"] = total_stats[1] if total_stats else 0
        report_data["퇴직근로자수"] = total_stats[2] if total_stats else 0

        # 2. 건강검진 실시 현황
        health_check_query = db.session.execute(
            text(
                """
            SELECT 
                COUNT(*) as total_checks,
                COUNT(CASE WHEN check_type = '일반' THEN 1 END) as general_checks,
                COUNT(CASE WHEN check_type = '특수' THEN 1 END) as special_checks,
                COUNT(CASE WHEN result = '정상' THEN 1 END) as normal_results,
                COUNT(CASE WHEN result = '관찰필요' THEN 1 END) as observation_needed,
                COUNT(CASE WHEN result = '치료필요' THEN 1 END) as treatment_needed
            FROM safework_health_checks 
            WHERE YEAR(check_date) = :year
        """
            ),
            {"year": year},
        )

        health_stats = health_check_query.fetchone()
        report_data["건강검진_총실시수"] = health_stats[0] if health_stats else 0
        report_data["일반검진수"] = health_stats[1] if health_stats else 0
        report_data["특수검진수"] = health_stats[2] if health_stats else 0
        report_data["정상판정수"] = health_stats[3] if health_stats else 0
        report_data["관찰필요수"] = health_stats[4] if health_stats else 0
        report_data["치료필요수"] = health_stats[5] if health_stats else 0

        # 3. 부서별 건강검진 현황
        dept_stats_query = db.session.execute(
            text(
                """
            SELECT 
                sw.department,
                COUNT(DISTINCT sw.id) as total_workers,
                COUNT(shc.id) as completed_checks,
                ROUND(COUNT(shc.id) * 100.0 / COUNT(DISTINCT sw.id), 1) as completion_rate
            FROM safework_workers sw
            LEFT JOIN safework_health_checks shc ON sw.id = shc.worker_id AND YEAR(shc.check_date) = :year
            WHERE sw.is_active = 1
            GROUP BY sw.department
            ORDER BY completion_rate DESC
        """
            ),
            {"year": year},
        )

        dept_data = []
        for row in dept_stats_query:
            dept_data.append(
                {
                    "부서명": row[0] or "미지정",
                    "전체인원": row[1] or 0,
                    "검진완료": row[2] or 0,
                    "완료율(%)": row[3] or 0,
                }
            )
        report_data["부서별현황"] = dept_data

        # 4. 월별 건강검진 추이
        monthly_query = db.session.execute(
            text(
                """
            SELECT 
                MONTH(check_date) as month,
                COUNT(*) as check_count
            FROM safework_health_checks
            WHERE YEAR(check_date) = :year
            GROUP BY MONTH(check_date)
            ORDER BY month
        """
            ),
            {"year": year},
        )

        monthly_data = []
        for row in monthly_query:
            monthly_data.append({"월": f"{row[0]}월", "검진수": row[1]})
        report_data["월별검진현황"] = monthly_data

        # 5. 의무실 방문 통계
        visit_stats_query = db.session.execute(
            text(
                """
            SELECT 
                COUNT(*) as total_visits,
                COUNT(CASE WHEN follow_up_needed = 1 THEN 1 END) as followup_needed,
                COUNT(CASE WHEN YEAR(visit_date) = :year AND MONTH(visit_date) = MONTH(CURDATE()) THEN 1 END) as this_month_visits
            FROM safework_medical_visits
            WHERE YEAR(visit_date) = :year
        """
            ),
            {"year": year},
        )

        visit_stats = visit_stats_query.fetchone()
        report_data["의무실방문_총횟수"] = visit_stats[0] if visit_stats else 0
        report_data["추적관찰필요"] = visit_stats[1] if visit_stats else 0
        report_data["이달방문횟수"] = visit_stats[2] if visit_stats else 0

        # JSON 형태로 응답
        if format_type == "json":
            return jsonify(
                {
                    "success": True,
                    "report_year": year,
                    "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "data": report_data,
                }
            )

        # Excel 파일로 응답
        elif format_type == "excel":
            return generate_excel_report(report_data, year)

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@safework_reports_bp.route("/medication-inventory")
@admin_required
def medication_inventory_report():
    """의약품 재고 현황 보고서"""
    try:
        format_type = request.args.get("format", "json")

        # 의약품 재고 상세 데이터
        inventory_query = db.session.execute(
            text(
                """
            SELECT 
                name, category, unit, current_stock, minimum_stock,
                expiry_date, supplier, price_per_unit, last_purchase_date,
                CASE 
                    WHEN current_stock = 0 THEN '재고없음'
                    WHEN current_stock <= minimum_stock THEN '재고부족'
                    ELSE '정상'
                END as stock_status,
                CASE 
                    WHEN expiry_date <= CURDATE() THEN '만료'
                    WHEN expiry_date <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN '만료임박'
                    ELSE '정상'
                END as expiry_status,
                COALESCE(current_stock * price_per_unit, 0) as total_value
            FROM safework_medications
            ORDER BY 
                CASE 
                    WHEN current_stock = 0 THEN 1
                    WHEN current_stock <= minimum_stock THEN 2
                    ELSE 3
                END,
                expiry_date ASC
        """
            )
        )

        medications = []
        total_value = 0
        alerts = {"재고없음": 0, "재고부족": 0, "만료": 0, "만료임박": 0}

        for row in inventory_query:
            med_data = {
                "의약품명": row[0],
                "분류": row[1] or "기타",
                "단위": row[2],
                "현재재고": row[3],
                "최소재고": row[4],
                "유효기간": row[5].strftime("%Y-%m-%d") if row[5] else "",
                "공급업체": row[6] or "",
                "단가": row[7] or 0,
                "최근입고일": row[8].strftime("%Y-%m-%d") if row[8] else "",
                "재고상태": row[9],
                "유효기간상태": row[10],
                "재고가치": row[11],
            }
            medications.append(med_data)
            total_value += row[11] or 0

            # 알림 카운트
            if row[9] in alerts:
                alerts[row[9]] += 1
            if row[10] in alerts:
                alerts[row[10]] += 1

        report_data = {
            "총의약품종수": len(medications),
            "총재고가치": total_value,
            "알림현황": alerts,
            "의약품목록": medications,
        }

        if format_type == "json":
            return jsonify(
                {
                    "success": True,
                    "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "data": report_data,
                }
            )
        elif format_type == "excel":
            return generate_medication_excel_report(medications, alerts, total_value)

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@safework_reports_bp.route("/worker-health-profile/<int:worker_id>")
@admin_required
def worker_health_profile(worker_id):
    """근로자 개인 건강 프로필 보고서"""
    try:
        # 근로자 기본 정보
        worker_query = db.session.execute(
            text(
                """
            SELECT employee_number, name, department, position, 
                   birth_date, hire_date, blood_type, phone, email
            FROM safework_workers 
            WHERE id = :worker_id
        """
            ),
            {"worker_id": worker_id},
        )

        worker_info = worker_query.fetchone()
        if not worker_info:
            return jsonify({"success": False, "error": "근로자를 찾을 수 없습니다."}), 404

        # 건강검진 이력
        health_checks_query = db.session.execute(
            text(
                """
            SELECT check_date, check_type, hospital, result, 
                   blood_pressure, bmi, findings, recommendations
            FROM safework_health_checks 
            WHERE worker_id = :worker_id 
            ORDER BY check_date DESC
            LIMIT 10
        """
            ),
            {"worker_id": worker_id},
        )

        health_checks = []
        for row in health_checks_query:
            health_checks.append(
                {
                    "검진일": row[0].strftime("%Y-%m-%d") if row[0] else "",
                    "검진구분": row[1] or "",
                    "검진기관": row[2] or "",
                    "판정": row[3] or "",
                    "혈압": row[4] or "",
                    "BMI": row[5] or "",
                    "소견": row[6] or "",
                    "권고사항": row[7] or "",
                }
            )

        # 의무실 방문 이력
        medical_visits_query = db.session.execute(
            text(
                """
            SELECT visit_date, chief_complaint, diagnosis, treatment, 
                   follow_up_needed, nurse_name
            FROM safework_medical_visits 
            WHERE worker_id = :worker_id 
            ORDER BY visit_date DESC
            LIMIT 20
        """
            ),
            {"worker_id": worker_id},
        )

        medical_visits = []
        for row in medical_visits_query:
            medical_visits.append(
                {
                    "방문일시": row[0].strftime("%Y-%m-%d %H:%M") if row[0] else "",
                    "주호소": row[1] or "",
                    "진단": row[2] or "",
                    "치료": row[3] or "",
                    "추적관찰필요": "예" if row[4] else "아니요",
                    "담당간호사": row[5] or "",
                }
            )

        # 나이 계산
        age = None
        if worker_info[4]:  # birth_date
            birth_date = (
                worker_info[4]
                if isinstance(worker_info[4], date)
                else datetime.strptime(worker_info[4], "%Y-%m-%d").date()
            )
            today = date.today()
            age = (
                today.year
                - birth_date.year
                - ((today.month, today.day) < (birth_date.month, birth_date.day))
            )

        profile_data = {
            "근로자정보": {
                "사번": worker_info[0],
                "성명": worker_info[1],
                "부서": worker_info[2] or "미지정",
                "직책": worker_info[3] or "사원",
                "생년월일": worker_info[4].strftime("%Y-%m-%d") if worker_info[4] else "",
                "나이": age,
                "입사일": worker_info[5].strftime("%Y-%m-%d") if worker_info[5] else "",
                "혈액형": worker_info[6] or "",
                "연락처": worker_info[7] or "",
                "이메일": worker_info[8] or "",
            },
            "건강검진이력": health_checks,
            "의무실방문이력": medical_visits,
            "통계": {
                "총건강검진수": len(health_checks),
                "총의무실방문수": len(medical_visits),
                "최근검진일": health_checks[0]["검진일"] if health_checks else "없음",
                "최근방문일": medical_visits[0]["방문일시"] if medical_visits else "없음",
            },
        }

        return jsonify(
            {
                "success": True,
                "worker_id": worker_id,
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "data": profile_data,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def generate_excel_report(report_data, year):
    """Excel 형태의 건강검진 보고서 생성"""
    output = io.BytesIO()
    workbook = Workbook()

    # 첫 번째 시트 - 종합 현황
    ws = workbook.active
    ws.title = "종합현황"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 제목
    ws["A1"] = f"SafeWork 건강검진 현황 보고서 ({year}년)"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    # 기본 통계
    ws["A3"] = "구분"
    ws["B3"] = "항목"
    ws["C3"] = "수치"
    ws["D3"] = "비고"

    for col in ["A3", "B3", "C3", "D3"]:
        ws[col].font = header_font
        ws[col].fill = header_fill
        ws[col].alignment = center_alignment

    row = 4
    basic_stats = [
        ("근로자 현황", "총 근로자 수", report_data.get("총근로자수", 0), "명"),
        ("", "재직 근로자 수", report_data.get("재직근로자수", 0), "명"),
        ("건강검진 현황", "총 검진 수", report_data.get("건강검진_총실시수", 0), "건"),
        ("", "일반검진", report_data.get("일반검진수", 0), "건"),
        ("", "특수검진", report_data.get("특수검진수", 0), "건"),
        ("판정 결과", "정상", report_data.get("정상판정수", 0), "건"),
        ("", "관찰필요", report_data.get("관찰필요수", 0), "건"),
        ("", "치료필요", report_data.get("치료필요수", 0), "건"),
    ]

    for stat in basic_stats:
        ws[f"A{row}"] = stat[0]
        ws[f"B{row}"] = stat[1]
        ws[f"C{row}"] = stat[2]
        ws[f"D{row}"] = stat[3]
        row += 1

    # 부서별 현황 시트
    if report_data.get("부서별현황"):
        ws2 = workbook.create_sheet(title="부서별현황")

        ws2["A1"] = "부서별 건강검진 현황"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2.merge_cells("A1:D1")

        headers = ["부서명", "전체인원", "검진완료", "완료율(%)"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        for row, dept in enumerate(report_data["부서별현황"], 4):
            ws2.cell(row=row, column=1, value=dept["부서명"])
            ws2.cell(row=row, column=2, value=dept["전체인원"])
            ws2.cell(row=row, column=3, value=dept["검진완료"])
            ws2.cell(row=row, column=4, value=dept["완료율(%)"])

    # 열 너비 자동 조정
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    filename = (
        f"SafeWork_건강검진보고서_{year}년_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def generate_medication_excel_report(medications, alerts, total_value):
    """의약품 재고 Excel 보고서 생성"""
    output = io.BytesIO()
    workbook = Workbook()
    ws = workbook.active
    ws.title = "의약품재고현황"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="28a745", end_color="28a745", fill_type="solid"
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 제목
    ws["A1"] = f"SafeWork 의약품 재고 현황 보고서"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:L1")

    # 요약 정보
    ws["A3"] = f"보고서 생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A4"] = f"총 의약품 종수: {len(medications)}종"
    ws["A5"] = f"총 재고 가치: {total_value:,.0f}원"
    ws["A6"] = f"재고 부족: {alerts.get('재고부족', 0)}종, 재고 없음: {alerts.get('재고없음', 0)}종"

    # 헤더
    headers = [
        "의약품명",
        "분류",
        "단위",
        "현재재고",
        "최소재고",
        "유효기간",
        "공급업체",
        "단가",
        "최근입고일",
        "재고상태",
        "유효기간상태",
        "재고가치",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # 데이터
    for row, med in enumerate(medications, 9):
        ws.cell(row=row, column=1, value=med["의약품명"])
        ws.cell(row=row, column=2, value=med["분류"])
        ws.cell(row=row, column=3, value=med["단위"])
        ws.cell(row=row, column=4, value=med["현재재고"])
        ws.cell(row=row, column=5, value=med["최소재고"])
        ws.cell(row=row, column=6, value=med["유효기간"])
        ws.cell(row=row, column=7, value=med["공급업체"])
        ws.cell(row=row, column=8, value=med["단가"])
        ws.cell(row=row, column=9, value=med["최근입고일"])
        ws.cell(row=row, column=10, value=med["재고상태"])
        ws.cell(row=row, column=11, value=med["유효기간상태"])
        ws.cell(row=row, column=12, value=med["재고가치"])

        # 재고 부족 행 강조
        if med["재고상태"] in ["재고없음", "재고부족"]:
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="ffe6e6", end_color="ffe6e6", fill_type="solid"
                )

    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)

    filename = f"SafeWork_의약품재고현황_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
